"""
Bulk import for tournaments via Excel — lets a pro_plus player upload one
.xlsx file containing the tournament, its teams, players/rosters, and
match fixtures, instead of entering everything one by one in the UI.
"""
import io
from datetime import datetime, date

from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST, require_GET
from django.db import transaction

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from tournaments.models import TournamentDetails, StartTournament
from teams.models import TeamDetails, PlayerDetails, TournamentTeam, TournamentRoster
from matches.models import CreateMatch

from subscriptions.decorators import require_plan, _is_privileged


VALID_TOURNAMENT_TYPES = {c[0] for c in TournamentDetails.TOURNAMENT_TYPE}
VALID_ROLES = {'BATSMAN', 'BOWLER', 'ALLROUNDER', 'WICKETKEEPER'}

HEADER_FILL = PatternFill(start_color='FF1F2937', end_color='FF1F2937', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFFFF')
EXAMPLE_FONT = Font(italic=True, color='FF6B7280')


# ─────────────────────────────────────────────────────────────────
#  TEMPLATE DOWNLOAD
# ─────────────────────────────────────────────────────────────────
@require_GET
@require_plan('pro_plus')
def bulk_import_template(request):
    wb = Workbook()

    # ---- Sheet 1: Tournament (single row of settings) ----
    ws = wb.active
    ws.title = "Tournament"
    headers = ["tournament_name", "tournament_type", "start_date", "end_date",
               "number_of_overs", "venue"]
    ws.append(headers)
    example = ["Summer Cup 2027", "OPEN_GROUND", "2027-06-01", "2027-06-30", 20,
               "Community Ground, Pune"]
    ws.append(example)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        ws.cell(row=2, column=col).font = EXAMPLE_FONT
        ws.column_dimensions[get_column_letter(col)].width = 22
    ws.cell(row=4, column=1, value=(
        "Fill row 2 with your real tournament. tournament_type must be exactly one of: "
        + ", ".join(sorted(VALID_TOURNAMENT_TYPES))
    )).font = Font(italic=True, size=9, color='FF9CA3AF')

    # ---- Sheet 2: Teams ----
    ws2 = wb.create_sheet("Teams")
    ws2.append(["team_name"])
    for name in ["Team Alpha", "Team Beta"]:
        ws2.append([name])
    ws2.cell(row=1, column=1).fill = HEADER_FILL
    ws2.cell(row=1, column=1).font = HEADER_FONT
    for r in (2, 3):
        ws2.cell(row=r, column=1).font = EXAMPLE_FONT
    ws2.column_dimensions['A'].width = 28

    # ---- Sheet 3: Players (also assigns each player to a team = roster) ----
    ws3 = wb.create_sheet("Players")
    p_headers = ["player_name", "mobile_number", "team_name", "role",
                 "is_captain", "is_vice_captain", "jersey_number"]
    ws3.append(p_headers)
    ws3.append(["Rahul Sharma", "9800000001", "Team Alpha", "BATSMAN", "YES", "NO", 7])
    ws3.append(["Amit Verma", "9800000002", "Team Alpha", "BOWLER", "NO", "NO", 11])
    ws3.append(["Vikram Singh", "9800000003", "Team Beta", "ALLROUNDER", "YES", "NO", 4])
    for col in range(1, len(p_headers) + 1):
        ws3.cell(row=1, column=col).fill = HEADER_FILL
        ws3.cell(row=1, column=col).font = HEADER_FONT
        ws3.column_dimensions[get_column_letter(col)].width = 18
        for r in (2, 3, 4):
            ws3.cell(row=r, column=col).font = EXAMPLE_FONT

    # ---- Sheet 4: Matches ----
    ws4 = wb.create_sheet("Matches")
    m_headers = ["team1_name", "team2_name", "match_date", "venue"]
    ws4.append(m_headers)
    ws4.append(["Team Alpha", "Team Beta", "2027-06-05", "Community Ground, Pune"])
    for col in range(1, len(m_headers) + 1):
        ws4.cell(row=1, column=col).fill = HEADER_FILL
        ws4.cell(row=1, column=col).font = HEADER_FONT
        ws4.column_dimensions[get_column_letter(col)].width = 22
        ws4.cell(row=2, column=col).font = EXAMPLE_FONT

    # ---- Sheet 5: Instructions (kept off the data sheets so a leftover note
    # row is never mistaken for a real team/player/match during import) ----
    ws5 = wb.create_sheet("Instructions")
    notes = [
        "Tournament sheet: fill row 2 with your real tournament. tournament_type must be exactly one of: "
        + ", ".join(sorted(VALID_TOURNAMENT_TYPES)),
        "Teams sheet: one row per team. Add as many rows as you need. Delete the example rows first.",
        "Players sheet: role must be one of: " + ", ".join(sorted(VALID_ROLES)) +
        ". is_captain/is_vice_captain: YES or NO. mobile_number must be unique per player. "
        "team_name must exactly match a name from the Teams sheet.",
        "Matches sheet: team1_name/team2_name must exactly match names from the Teams sheet. "
        "match_date format: YYYY-MM-DD.",
        "On every sheet: stop your data at the last real row and leave everything below it empty — "
        "the first blank row is treated as the end of the list.",
    ]
    for i, note in enumerate(notes, start=1):
        ws5.cell(row=i, column=1, value=note).font = Font(size=10, color='FF374151')
    ws5.column_dimensions['A'].width = 110

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = 'attachment; filename="tournament_import_template.xlsx"'
    return resp


# ─────────────────────────────────────────────────────────────────
#  UPLOAD + VALIDATE + IMPORT
# ─────────────────────────────────────────────────────────────────
def _parse_date(val, field_label, errors):
    if val is None or val == '':
        errors.append(f"{field_label} is required")
        return None
    if isinstance(val, (datetime, date)):
        return val if isinstance(val, date) and not isinstance(val, datetime) else val.date()
    try:
        return datetime.strptime(str(val).strip(), "%Y-%m-%d").date()
    except ValueError:
        errors.append(f"{field_label} '{val}' is not a valid date (use YYYY-MM-DD)")
        return None


def _bool_cell(val):
    return str(val).strip().upper() in ('YES', 'TRUE', '1', 'Y')


@require_POST
@require_plan('pro_plus')
def bulk_import_upload(request):
    f = request.FILES.get('excel_file')
    if not f:
        return JsonResponse({'success': False, 'errors': ["No file was uploaded."]}, status=400)

    if not f.name.lower().endswith('.xlsx'):
        return JsonResponse({'success': False, 'errors': [
            "Please upload a .xlsx file (the format the template downloads as). "
            "Older .xls or .csv files are not supported."
        ]}, status=400)

    try:
        wb = load_workbook(io.BytesIO(f.read()), data_only=True)
    except Exception as e:
        return JsonResponse({'success': False, 'errors': [
            f"Could not open this file as an Excel workbook: {e}"
        ]}, status=400)

    errors = []
    required_sheets = ['Tournament', 'Teams', 'Players', 'Matches']
    for s in required_sheets:
        if s not in wb.sheetnames:
            errors.append(f"Missing required sheet: '{s}'. Use the downloaded template as your starting point.")
    if errors:
        return JsonResponse({'success': False, 'errors': errors}, status=400)

    # ---- Parse Tournament sheet (row 2) ----
    t_ws = wb['Tournament']
    t_row = [c.value for c in t_ws[2]] if t_ws.max_row >= 2 else []
    while len(t_row) < 6:
        t_row.append(None)
    tournament_name, tournament_type, start_date_raw, end_date_raw, overs_raw, venue = t_row[:6]

    if not tournament_name or not str(tournament_name).strip():
        errors.append("Tournament sheet: tournament_name is required in row 2.")
    if tournament_type not in VALID_TOURNAMENT_TYPES:
        errors.append(
            f"Tournament sheet: tournament_type '{tournament_type}' is invalid. "
            f"Must be one of: {', '.join(sorted(VALID_TOURNAMENT_TYPES))}"
        )
    start_date_val = _parse_date(start_date_raw, "Tournament sheet: start_date", errors)
    end_date_val = _parse_date(end_date_raw, "Tournament sheet: end_date", errors)
    try:
        overs_val = int(overs_raw)
        if overs_val <= 0:
            raise ValueError
    except (TypeError, ValueError):
        errors.append("Tournament sheet: number_of_overs must be a positive whole number.")
        overs_val = None
    if TournamentDetails.objects.filter(tournament_name=str(tournament_name).strip()).exists():
        errors.append(
            f"A tournament named '{tournament_name}' already exists. "
            "Use a different name to avoid mixing up data."
        )

    # ---- Parse Teams sheet ----
    teams_ws = wb['Teams']
    team_names = []
    for row in teams_ws.iter_rows(min_row=2, values_only=True):
        # A blank row marks the end of real data — anything after it (like the
        # instructional note the template leaves below the example rows) is not
        # a team and must not be validated as one.
        if not row or not row[0]:
            break
        name = str(row[0]).strip()
        if name in team_names:
            errors.append(f"Teams sheet: duplicate team name '{name}'.")
        team_names.append(name)
    if len(team_names) < 2:
        errors.append("Teams sheet: at least 2 teams are required.")

    # ---- Parse Players sheet ----
    players_ws = wb['Players']
    player_rows = []
    seen_mobiles = set()
    for i, row in enumerate(players_ws.iter_rows(min_row=2, values_only=True), start=2):
        # First blank row = end of real data (see Teams sheet comment above).
        if not row or not any(row):
            break
        row = list(row) + [None] * (7 - len(row))
        p_name, mobile, team_name, role, is_cap, is_vc, jersey = row[:7]

        if not p_name or not str(p_name).strip():
            errors.append(f"Players sheet row {i}: player_name is required.")
            continue
        if not mobile or not str(mobile).strip():
            errors.append(f"Players sheet row {i} ({p_name}): mobile_number is required.")
            continue
        mobile = str(mobile).strip()
        if mobile in seen_mobiles:
            errors.append(f"Players sheet row {i} ({p_name}): mobile_number '{mobile}' is duplicated in this file.")
        seen_mobiles.add(mobile)

        if not team_name or str(team_name).strip() not in team_names:
            errors.append(
                f"Players sheet row {i} ({p_name}): team_name '{team_name}' doesn't match any "
                f"team in the Teams sheet."
            )
            continue
        role_val = str(role).strip().upper() if role else ''
        if role_val not in VALID_ROLES:
            errors.append(
                f"Players sheet row {i} ({p_name}): role '{role}' is invalid. "
                f"Must be one of: {', '.join(sorted(VALID_ROLES))}"
            )
            continue
        try:
            jersey_val = int(jersey) if jersey not in (None, '') else None
        except (TypeError, ValueError):
            errors.append(f"Players sheet row {i} ({p_name}): jersey_number must be a whole number.")
            jersey_val = None

        player_rows.append({
            'name': str(p_name).strip(),
            'mobile': mobile,
            'team': str(team_name).strip(),
            'role': role_val,
            'is_captain': _bool_cell(is_cap),
            'is_vice_captain': _bool_cell(is_vc),
            'jersey': jersey_val,
        })
    if not player_rows and not errors:
        errors.append("Players sheet: no players found. Add at least one player per team.")

    # Cross-check: a mobile number already used by an existing player is fine (we reuse it),
    # but flag if the SAME mobile appears assigned to two different names within this file.
    for pr in player_rows:
        existing = PlayerDetails.objects.filter(mobile_number=pr['mobile']).first()
        if existing and existing.player_name.strip().lower() != pr['name'].strip().lower():
            errors.append(
                f"Players sheet: mobile_number '{pr['mobile']}' is already registered to "
                f"'{existing.player_name}' in the system, but this file lists it as '{pr['name']}'. "
                f"Fix the mobile number or the name so they match."
            )

    # ---- Parse Matches sheet ----
    matches_ws = wb['Matches']
    match_rows = []
    for i, row in enumerate(matches_ws.iter_rows(min_row=2, values_only=True), start=2):
        # First blank row = end of real data (see Teams sheet comment above).
        if not row or not any(row):
            break
        row = list(row) + [None] * (4 - len(row))
        t1, t2, m_date_raw, m_venue = row[:4]

        if not t1 or str(t1).strip() not in team_names:
            errors.append(f"Matches sheet row {i}: team1_name '{t1}' doesn't match any team in the Teams sheet.")
            continue
        if not t2 or str(t2).strip() not in team_names:
            errors.append(f"Matches sheet row {i}: team2_name '{t2}' doesn't match any team in the Teams sheet.")
            continue
        if str(t1).strip() == str(t2).strip():
            errors.append(f"Matches sheet row {i}: a team can't play itself ({t1}).")
            continue
        m_date_val = _parse_date(m_date_raw, f"Matches sheet row {i}: match_date", errors)
        if not m_venue or not str(m_venue).strip():
            errors.append(f"Matches sheet row {i}: venue is required.")
            continue
        if m_date_val is None:
            continue

        match_rows.append({
            'team1': str(t1).strip(),
            'team2': str(t2).strip(),
            'date': m_date_val,
            'venue': str(m_venue).strip(),
        })

    if errors:
        return JsonResponse({'success': False, 'errors': errors}, status=400)

    # ---- Everything validated — commit to the database ----
    try:
        with transaction.atomic():
            tournament = TournamentDetails.objects.create(
                tournament_name=str(tournament_name).strip(),
                tournament_type=tournament_type,
                start_date=start_date_val,
                end_date=end_date_val,
                number_of_overs=overs_val,
                number_of_teams=len(team_names),
                venue=str(venue).strip() if venue else '',
            )

            # Owner: whoever is uploading (player) — falls back to admin-created if staff
            player_id = request.session.get('player_id')
            if player_id and player_id != 'guest' and not _is_privileged(request):
                tournament.created_by_player_id = player_id
                tournament.save(update_fields=['created_by_player_id'])
            elif request.user.is_authenticated and (request.user.is_staff or request.user.is_superuser):
                tournament.created_by_admin_id = request.user.id
                tournament.save(update_fields=['created_by_admin_id'])

            StartTournament.objects.get_or_create(tournament=tournament, defaults={'is_started': True})

            team_objs = {}
            for name in team_names:
                team, _ = TeamDetails.objects.get_or_create(team_name=name)
                team_objs[name] = team
                TournamentTeam.objects.get_or_create(tournament=tournament, team=team)

            player_objs = {}
            for pr in player_rows:
                player, _ = PlayerDetails.objects.get_or_create(
                    mobile_number=pr['mobile'],
                    defaults={'player_name': pr['name']},
                )
                player_objs[pr['mobile']] = player

                tt = TournamentTeam.objects.get(tournament=tournament, team=team_objs[pr['team']])
                TournamentRoster.objects.get_or_create(
                    tournament_team=tt,
                    tournament=tournament,
                    player=player,
                    defaults={
                        'role': pr['role'],
                        'is_captain': pr['is_captain'],
                        'is_vice_captain': pr['is_vice_captain'],
                        'jersey_number': pr['jersey'],
                    },
                )

            for mr in match_rows:
                CreateMatch.objects.create(
                    tournament=tournament,
                    team1=team_objs[mr['team1']],
                    team2=team_objs[mr['team2']],
                    match_date=mr['date'],
                    venue=mr['venue'],
                )

    except Exception as e:
        return JsonResponse({'success': False, 'errors': [
            f"Import failed while saving to the database: {e}"
        ]}, status=500)

    return JsonResponse({
        'success': True,
        'tournament_id': tournament.id,
        'tournament_name': tournament.tournament_name,
        'teams_created': len(team_names),
        'players_created': len(player_rows),
        'matches_created': len(match_rows),
    })
